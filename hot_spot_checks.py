import win32api
from openpyxl import load_workbook
from datetime import datetime, timedelta


# Function: Datetime stamp to string date (DD/MM/YYYY)
def dateStampToStrDate(dateStamp):
    strDate = datetime.strftime(dateStamp, '%d/%m/%Y')
    return strDate


# Function: String date (DD/MM/YYYY) to Datetime stamp
def strDatetoDateStamp(strDate):
    datestamp = datetime.strptime(strDate, '%d/%m/%Y')
    return datestamp


# Function: Print File 'fn' from Path 'path'
def printFile(path, fn):
    win32api.ShellExecute(0, "print", path + fn, None, ".", 0)


# Function: Take date in file, add 7 days, save file.
def UpdateFile(path, fn, cellRef):
    wb = load_workbook(path + fn)
    ws = wb.get_sheet_by_name('Sheet1')
    strDate = input('Put in week ending data: ')
    ws.cell(coordinate=cellRef).value = strDate
    wb.save(path + fn)


firepath = 'U:\\rmarshall\\Yard Reports\\Fire Hot Spots\\'
shedpath = 'U:\\rmarshall\\Yard Reports\\\Waste Shed Checks\\'
firename = 'Fire Hot Spots & Washing Plant Check Sheet.xlsx'
shedname = 'Waste Shed Checks.xlsx'

UpdateFile(firepath, firename, 'C1')
UpdateFile(shedpath, shedname, 'C2')

printFile(firepath, firename)  # print Fire Hot Spots & Wash Plant Checks sheet (1st Copy)
printFile(firepath, firename)  # print Fire Hot Spots & Wash Plant Checks sheet (2nd Copy)
printFile(shedpath, shedname)  # print Waste Shed Checks sheet
printFile(shedpath, shedname)  # print Waste Shed Checks sheet